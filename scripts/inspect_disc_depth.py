from pathlib import Path
from openpyxl import load_workbook

workbook = load_workbook(Path("/home/ubuntu/upload/Database_SaaS_BHNT(1).xlsx"), data_only=True)
sheet = workbook["10_Trạm Đăng Kiểm Năng Lực"]
print({"title": sheet.title, "max_row": sheet.max_row, "max_column": sheet.max_column, "dimension": sheet.calculate_dimension()})
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
    if any(value not in (None, "") for value in row):
        print([value for value in row])
