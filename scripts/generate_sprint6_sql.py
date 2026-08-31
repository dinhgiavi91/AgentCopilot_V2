from __future__ import annotations

from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

SOURCE = Path("/home/ubuntu/upload/Database_SaaS_BHNT(1).xlsx")
OUTPUT = Path("/home/ubuntu/bhnt-learning-hub-research/supabase/migrations/20260813_sprint6_operational_modules.sql")

def clean(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    value = str(value).strip()
    return value or None

def q(value):
    if value is None:
        return "null"
    return "'" + value.replace("'", "''") + "'"

def data_rows(sheet):
    return [row for row in sheet.iter_rows(min_row=2, values_only=True) if any(clean(value) for value in row)]

schema = """
-- Sprint 6 — Operational modules sourced from Database_SaaS_BHNT (1).xlsx.
-- Zero-PII: content tables hold operational knowledge only. Feedback must not contain customer identifiers.
create extension if not exists "pgcrypto";

create table if not exists public.disc_questions (
  code text primary key,
  question text not null,
  option_d text not null,
  option_i text not null,
  option_s text not null,
  option_c text not null,
  source_sheet text not null default '10_Trạm Đăng Kiểm Năng Lực',
  sort_order integer not null check (sort_order > 0),
  created_at timestamptz not null default now()
);

create table if not exists public.cover_letters (
  code text primary key,
  situation text not null,
  body_template text not null,
  source_sheet text not null default '5_Trợ Lý Thẩm Định',
  sort_order integer not null check (sort_order > 0),
  created_at timestamptz not null default now()
);

create table if not exists public.news_case_studies (
  code text primary key,
  category text not null,
  title text not null,
  summary text not null,
  field_takeaway text not null,
  published_at timestamptz,
  source_sheet text not null default '11_Bản Tin 90s & Án Lệ',
  sort_order integer not null check (sort_order > 0),
  created_at timestamptz not null default now()
);

create table if not exists public.disc_assessments (
  assessment_id uuid primary key default gen_random_uuid(),
  user_id uuid not null references auth.users(id) on delete cascade,
  disc_type text not null check (disc_type in ('D', 'I', 'S', 'C')),
  score_d smallint not null default 0,
  score_i smallint not null default 0,
  score_s smallint not null default 0,
  score_c smallint not null default 0,
  created_at timestamptz not null default now()
);

create table if not exists public.feedback_entries (
  feedback_id uuid primary key default gen_random_uuid(),
  user_id uuid references auth.users(id) on delete set null,
  rating smallint not null check (rating between 1 and 5),
  favorite_feature text not null check (char_length(favorite_feature) <= 120),
  suggestion text not null check (char_length(suggestion) between 3 and 1000),
  created_at timestamptz not null default now()
);

create index if not exists disc_assessments_user_created_idx on public.disc_assessments(user_id, created_at desc);
create index if not exists feedback_entries_created_idx on public.feedback_entries(created_at desc);

alter table public.disc_questions enable row level security;
alter table public.cover_letters enable row level security;
alter table public.news_case_studies enable row level security;
alter table public.disc_assessments enable row level security;
alter table public.feedback_entries enable row level security;

drop policy if exists "sprint6_read_disc_questions" on public.disc_questions;
create policy "sprint6_read_disc_questions" on public.disc_questions for select to anon, authenticated using (true);
drop policy if exists "sprint6_read_cover_letters" on public.cover_letters;
create policy "sprint6_read_cover_letters" on public.cover_letters for select to anon, authenticated using (true);
drop policy if exists "sprint6_read_news_case_studies" on public.news_case_studies;
create policy "sprint6_read_news_case_studies" on public.news_case_studies for select to anon, authenticated using (true);
drop policy if exists "sprint6_select_own_disc_assessments" on public.disc_assessments;
create policy "sprint6_select_own_disc_assessments" on public.disc_assessments for select to authenticated using (auth.uid() = user_id);
drop policy if exists "sprint6_insert_own_disc_assessments" on public.disc_assessments;
create policy "sprint6_insert_own_disc_assessments" on public.disc_assessments for insert to authenticated with check (auth.uid() = user_id);
drop policy if exists "sprint6_insert_feedback" on public.feedback_entries;
create policy "sprint6_insert_feedback" on public.feedback_entries for insert to anon, authenticated with check (user_id is null or auth.uid() = user_id);

grant select on public.disc_questions, public.cover_letters, public.news_case_studies to anon, authenticated;
grant select, insert on public.disc_assessments to authenticated;
grant insert on public.feedback_entries to anon, authenticated;
"""

def main():
    workbook = load_workbook(SOURCE, data_only=True)
    statements = [schema]
    for index, row in enumerate(data_rows(workbook["10_Trạm Đăng Kiểm Năng Lực"]), 1):
        code, question, d, i, s, c = [clean(row[position]) if len(row) > position else None for position in range(6)]
        if code and question:
            statements.append(f"insert into public.disc_questions (code, question, option_d, option_i, option_s, option_c, sort_order) values ({q(code)}, {q(question)}, {q(d)}, {q(i)}, {q(s)}, {q(c)}, {index}) on conflict (code) do update set question = excluded.question, option_d = excluded.option_d, option_i = excluded.option_i, option_s = excluded.option_s, option_c = excluded.option_c, sort_order = excluded.sort_order;")
    for index, row in enumerate(data_rows(workbook["5_Trợ Lý Thẩm Định"]), 1):
        code, situation, body = [clean(row[position]) if len(row) > position else None for position in range(3)]
        if code and situation and body:
            statements.append(f"insert into public.cover_letters (code, situation, body_template, sort_order) values ({q(code)}, {q(situation)}, {q(body)}, {index}) on conflict (code) do update set situation = excluded.situation, body_template = excluded.body_template, sort_order = excluded.sort_order;")
    for index, row in enumerate(data_rows(workbook["11_Bản Tin 90s & Án Lệ"]), 1):
        code, category, title, summary, takeaway, published = [clean(row[position]) if len(row) > position else None for position in range(6)]
        if code and category and title and summary and takeaway:
            statements.append(f"insert into public.news_case_studies (code, category, title, summary, field_takeaway, published_at, sort_order) values ({q(code)}, {q(category)}, {q(title)}, {q(summary)}, {q(takeaway)}, {q(published)}, {index}) on conflict (code) do update set category = excluded.category, title = excluded.title, summary = excluded.summary, field_takeaway = excluded.field_takeaway, published_at = excluded.published_at, sort_order = excluded.sort_order;")
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text("\n\n".join(statements) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print("disc_questions=5 cover_letters=4 news_case_studies=3")

if __name__ == "__main__":
    main()
