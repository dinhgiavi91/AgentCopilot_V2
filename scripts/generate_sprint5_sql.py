from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

SOURCE = Path("/home/ubuntu/upload/Database_SaaS_BHNT(1).xlsx")
OUTPUT = Path("/home/ubuntu/bhnt-learning-hub-research/supabase/migrations/20260813_sprint5_content_library.sql")

DDL = """-- Sprint 5: Content Library sourced from Database_SaaS_BHNT(1).xlsx
-- This migration is idempotent and safe to paste into Supabase SQL Editor.

create extension if not exists pgcrypto;

create table if not exists public.playbook_cards (
  code text primary key,
  skill_system text not null,
  required_level text not null default 'Rookie',
  situation text not null,
  mindset text not null,
  coaching_prompts text,
  is_pro boolean not null default false,
  source_sheet text not null default '3_Bảo Bối Thực Chiến',
  sort_order integer not null default 0,
  created_at timestamptz not null default now(),
  updated_at timestamptz not null default now()
);
alter table public.playbook_cards add column if not exists skill_system text;
alter table public.playbook_cards add column if not exists required_level text;
alter table public.playbook_cards add column if not exists situation text;
alter table public.playbook_cards add column if not exists mindset text;
alter table public.playbook_cards add column if not exists coaching_prompts text;
alter table public.playbook_cards add column if not exists is_pro boolean not null default false;
alter table public.playbook_cards add column if not exists source_sheet text;
alter table public.playbook_cards add column if not exists sort_order integer not null default 0;
alter table public.playbook_cards add column if not exists created_at timestamptz not null default now();
alter table public.playbook_cards add column if not exists updated_at timestamptz not null default now();

create table if not exists public.empathy_dictionary (
  code text primary key,
  legal_term text not null,
  empathy_translation text not null,
  source_sheet text not null default '4_Ngôn Ngữ Thấu Cảm',
  sort_order integer not null default 0,
  created_at timestamptz not null default now(),
  updated_at timestamptz not null default now()
);
alter table public.empathy_dictionary add column if not exists legal_term text;
alter table public.empathy_dictionary add column if not exists empathy_translation text;
alter table public.empathy_dictionary add column if not exists source_sheet text;
alter table public.empathy_dictionary add column if not exists sort_order integer not null default 0;
alter table public.empathy_dictionary add column if not exists created_at timestamptz not null default now();
alter table public.empathy_dictionary add column if not exists updated_at timestamptz not null default now();

create table if not exists public.leadership_compass (
  code text primary key,
  topic text not null,
  core_thinking text not null,
  source_sheet text not null default '6_La Bàn Lãnh Đạo',
  sort_order integer not null default 0,
  created_at timestamptz not null default now(),
  updated_at timestamptz not null default now()
);
alter table public.leadership_compass add column if not exists topic text;
alter table public.leadership_compass add column if not exists core_thinking text;
alter table public.leadership_compass add column if not exists source_sheet text;
alter table public.leadership_compass add column if not exists sort_order integer not null default 0;
alter table public.leadership_compass add column if not exists created_at timestamptz not null default now();
alter table public.leadership_compass add column if not exists updated_at timestamptz not null default now();

create table if not exists public.marketing_templates (
  code text primary key,
  category text not null,
  occasion text not null,
  message_template text not null,
  image_url text,
  source_sheet text not null default '13_Marketing 1 Chạm',
  sort_order integer not null default 0,
  created_at timestamptz not null default now(),
  updated_at timestamptz not null default now()
);
alter table public.marketing_templates add column if not exists category text;
alter table public.marketing_templates add column if not exists occasion text;
alter table public.marketing_templates add column if not exists message_template text;
alter table public.marketing_templates add column if not exists image_url text;
alter table public.marketing_templates add column if not exists source_sheet text;
alter table public.marketing_templates add column if not exists sort_order integer not null default 0;
alter table public.marketing_templates add column if not exists created_at timestamptz not null default now();
alter table public.marketing_templates add column if not exists updated_at timestamptz not null default now();

alter table public.playbook_cards enable row level security;
alter table public.empathy_dictionary enable row level security;
alter table public.leadership_compass enable row level security;
alter table public.marketing_templates enable row level security;

drop policy if exists "content_library_read_playbook" on public.playbook_cards;
drop policy if exists "content_library_read_empathy" on public.empathy_dictionary;
drop policy if exists "content_library_read_leadership" on public.leadership_compass;
drop policy if exists "content_library_read_marketing" on public.marketing_templates;
create policy "content_library_read_playbook" on public.playbook_cards for select to anon, authenticated using (true);
create policy "content_library_read_empathy" on public.empathy_dictionary for select to anon, authenticated using (true);
create policy "content_library_read_leadership" on public.leadership_compass for select to anon, authenticated using (true);
create policy "content_library_read_marketing" on public.marketing_templates for select to anon, authenticated using (true);
grant select on public.playbook_cards, public.empathy_dictionary, public.leadership_compass, public.marketing_templates to anon, authenticated;

"""

def cell(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None

def quote(value):
    if value is None:
        return "null"
    return "'" + value.replace("'", "''") + "'"

def upsert(table, columns, rows):
    result = []
    update_columns = [column for column in columns if column not in {"code", "created_at"}]
    for row in rows:
        values = ", ".join(serialize(value) for value in row)
        updates = ", ".join(f"{column} = excluded.{column}" for column in update_columns)
        result.append(f"insert into public.{table} ({', '.join(columns)}) values ({values}) on conflict (code) do update set {updates}, updated_at = now();")
    return "\n".join(result)

def serialize(value):
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    return quote(str(value))

def nonempty_rows(sheet):
    return [row for row in sheet.iter_rows(min_row=2, values_only=True) if any(cell(value) for value in row)]

def main():
    workbook = load_workbook(SOURCE, data_only=True)
    playbooks = []
    for index, row in enumerate(nonempty_rows(workbook["3_Bảo Bối Thực Chiến"]), 1):
        code, skill, level, situation, mindset, prompts = [cell(row[position]) if len(row) > position else None for position in range(6)]
        if not code or not situation or not mindset:
            continue
        playbooks.append((code, skill or "Bảo Bối", level or "Rookie", situation, mindset, prompts, (level or "").lower() != "rookie", "3_Bảo Bối Thực Chiến", index))

    empathy = []
    for index, row in enumerate(nonempty_rows(workbook["4_Ngôn Ngữ Thấu Cảm"]), 1):
        code, legal_term, translation = [cell(row[position]) if len(row) > position else None for position in range(3)]
        if code and legal_term and translation:
            empathy.append((code, legal_term, translation, "4_Ngôn Ngữ Thấu Cảm", index))

    leadership = []
    for index, row in enumerate(nonempty_rows(workbook["6_La Bàn Lãnh Đạo"]), 1):
        code, topic, thinking = [cell(row[position]) if len(row) > position else None for position in range(3)]
        if code and topic and thinking:
            leadership.append((code, topic, thinking, "6_La Bàn Lãnh Đạo", index))

    marketing = []
    for index, row in enumerate(nonempty_rows(workbook["13_Marketing 1 Chạm"]), 1):
        code, category, occasion, message, image_url = [cell(row[position]) if len(row) > position else None for position in range(5)]
        if code and category and occasion and message:
            marketing.append((code, category, occasion, message, image_url, "13_Marketing 1 Chạm", index))

    sections = [
        DDL,
        "-- Seed: 3_Bảo Bối Thực Chiến\n" + upsert("playbook_cards", ["code", "skill_system", "required_level", "situation", "mindset", "coaching_prompts", "is_pro", "source_sheet", "sort_order"], playbooks),
        "-- Seed: 4_Ngôn Ngữ Thấu Cảm\n" + upsert("empathy_dictionary", ["code", "legal_term", "empathy_translation", "source_sheet", "sort_order"], empathy),
        "-- Seed: 6_La Bàn Lãnh Đạo\n" + upsert("leadership_compass", ["code", "topic", "core_thinking", "source_sheet", "sort_order"], leadership),
        "-- Seed: 13_Marketing 1 Chạm\n" + upsert("marketing_templates", ["code", "category", "occasion", "message_template", "image_url", "source_sheet", "sort_order"], marketing),
        f"\n-- Expected seeded rows: playbook_cards={len(playbooks)}, empathy_dictionary={len(empathy)}, leadership_compass={len(leadership)}, marketing_templates={len(marketing)}\n",
    ]
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text("\n\n".join(sections), encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(f"playbook_cards={len(playbooks)} empathy_dictionary={len(empathy)} leadership_compass={len(leadership)} marketing_templates={len(marketing)}")

if __name__ == "__main__":
    main()
