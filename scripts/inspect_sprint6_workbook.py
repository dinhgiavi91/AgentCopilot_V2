from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook

source = Path("/home/ubuntu/upload/Database_SaaS_BHNT(1).xlsx")
output = Path("/home/ubuntu/bhnt-learning-hub-research/data/sprint6_workbook_inspection.json")

def normalize(value):
    return None if value is None else str(value).strip()

workbook = load_workbook(source, data_only=True)
result = {}
for sheet in workbook.worksheets:
    rows = [[normalize(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    non_empty = [row for row in rows if any(cell for cell in row)]
    result[sheet.title] = {
        "rows": len(non_empty),
        "headers": non_empty[0] if non_empty else [],
        "sample": non_empty[1:6] if len(non_empty) > 1 else [],
    }
output.parent.mkdir(parents=True, exist_ok=True)
output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
print(output)
