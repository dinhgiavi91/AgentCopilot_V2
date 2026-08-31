from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook

SOURCE = Path("/home/ubuntu/upload/Database_SaaS_BHNT(1).xlsx")
OUTPUT = Path("/home/ubuntu/bhnt-learning-hub-research/data/sprint5_seed.json")

def clean(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None

def rows(sheet):
    return [row for row in sheet.iter_rows(min_row=2, values_only=True) if any(clean(item) for item in row)]

def main():
    workbook = load_workbook(SOURCE, data_only=True)
    playbooks = []
    for index, row in enumerate(rows(workbook["3_Bảo Bối Thực Chiến"]), 1):
        values = [clean(row[position]) if len(row) > position else None for position in range(6)]
        code, skill_system, required_level, situation, mindset, coaching_prompts = values
        if code and situation and mindset:
            playbooks.append({"code": code, "skill_system": skill_system or "Bảo Bối", "required_level": required_level or "Rookie", "situation": situation, "mindset": mindset, "coaching_prompts": coaching_prompts, "is_pro": (required_level or "").lower() != "rookie", "source_sheet": "3_Bảo Bối Thực Chiến", "sort_order": index})
    empathy = []
    for index, row in enumerate(rows(workbook["4_Ngôn Ngữ Thấu Cảm"]), 1):
        code, legal_term, empathy_translation = [clean(row[position]) if len(row) > position else None for position in range(3)]
        if code and legal_term and empathy_translation:
            empathy.append({"code": code, "legal_term": legal_term, "empathy_translation": empathy_translation, "source_sheet": "4_Ngôn Ngữ Thấu Cảm", "sort_order": index})
    leadership = []
    for index, row in enumerate(rows(workbook["6_La Bàn Lãnh Đạo"]), 1):
        code, topic, core_thinking = [clean(row[position]) if len(row) > position else None for position in range(3)]
        if code and topic and core_thinking:
            leadership.append({"code": code, "topic": topic, "core_thinking": core_thinking, "source_sheet": "6_La Bàn Lãnh Đạo", "sort_order": index})
    marketing = []
    for index, row in enumerate(rows(workbook["13_Marketing 1 Chạm"]), 1):
        code, category, occasion, message_template, image_url = [clean(row[position]) if len(row) > position else None for position in range(5)]
        if code and category and occasion and message_template:
            marketing.append({"code": code, "category": category, "occasion": occasion, "message_template": message_template, "image_url": image_url, "source_sheet": "13_Marketing 1 Chạm", "sort_order": index})
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps({"playbook_cards": playbooks, "empathy_dictionary": empathy, "leadership_compass": leadership, "marketing_templates": marketing}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(f"playbook_cards={len(playbooks)} empathy_dictionary={len(empathy)} leadership_compass={len(leadership)} marketing_templates={len(marketing)}")

if __name__ == "__main__":
    main()
