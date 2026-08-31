from pathlib import Path
from openpyxl import load_workbook

source = Path("/home/ubuntu/upload/Database_SaaS_BHNT(1).xlsx")
workbook = load_workbook(source, data_only=True)

for sheet in workbook.worksheets:
    print(f"\n=== {sheet.title} ===")
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True):
        cleaned = ["" if value is None else str(value).replace("\n", " ↵ ") for value in row]
        print(" | ".join(cleaned))
