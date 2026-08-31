from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook

SOURCE = Path("/home/ubuntu/upload/Database_SaaS_BHNT(1).xlsx")
TARGETS = [
    "3_Bảo Bối Thực Chiến",
    "4_Ngôn Ngữ Thấu Cảm",
    "6_La Bàn Lãnh Đạo",
    "13_Marketing 1 Chạm",
]

def normalize(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value).strip()

def main():
    workbook = load_workbook(SOURCE, data_only=True)
    result = {"sheets": workbook.sheetnames, "targets": {}}
    for name in TARGETS:
        sheet = workbook[name]
        rows = [[normalize(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        nonempty = [row for row in rows if any(value not in (None, "") for value in row)]
        result["targets"][name] = {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "nonempty_row_count": len(nonempty),
            "rows": nonempty[:20],
        }
    output = Path("/home/ubuntu/bhnt-learning-hub-research/data/sprint5_workbook_inspection.json")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output)

if __name__ == "__main__":
    main()
