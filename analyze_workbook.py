from pathlib import Path
import json
import re
from openpyxl import load_workbook

SOURCE = Path('/home/ubuntu/upload/Database_SaaS_BHNT(1).xlsx')
OUT = Path('/home/ubuntu/bhnt-learning-hub-research/workbook_summary.json')

wb = load_workbook(SOURCE, data_only=True)
summary = {'source': str(SOURCE), 'sheets': [], 'derived': {}}

def safe_text(value):
    return '' if value is None else str(value).strip()

for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=True))
    headers = [safe_text(v) for v in rows[0]] if rows else []
    data = rows[1:] if len(rows) > 1 else []
    non_empty_data = [row for row in data if any(v is not None and safe_text(v) for v in row)]
    sheet = {
        'name': ws.title,
        'rows': len(data),
        'non_empty_rows': len(non_empty_data),
        'columns': len(headers),
        'headers': headers,
        'sample': [[safe_text(v) for v in row] for row in non_empty_data[:3]],
    }
    summary['sheets'].append(sheet)

profiles = next((s for s in summary['sheets'] if s['name'] == '1_Hồ Sơ Chiến Binh'), None)
if profiles:
    summary['derived']['profile_counts'] = {
        'total': profiles['non_empty_rows'],
        'role_fields': [h for h in profiles['headers'] if 'Phân Quyền' in h or 'Cấp Độ' in h or 'Trạng Thái' in h or 'Nhóm DISC' in h],
    }

radar = next((s for s in summary['sheets'] if s['name'] == '14_Radar Giữ Quân'), None)
if radar:
    summary['derived']['radar'] = {
        'total_cases': radar['non_empty_rows'],
        'status_field': next((h for h in radar['headers'] if 'Trạng Thái' in h), None),
    }

logs = next((s for s in summary['sheets'] if s['name'] == '2_Nhịp Đập Khách Hàng'), None)
if logs:
    summary['derived']['customer_pulse'] = {
        'total_logs': logs['non_empty_rows'],
        'pii_sensitive_headers': [h for h in logs['headers'] if any(term in h.lower() for term in ['ghi', 'khách', 'số', 'email', 'phone', 'cccd'])],
        'privacy_header': next((h for h in logs['headers'] if 'Riêng Tư' in h), None),
    }

summary['derived']['sheet_names'] = [s['name'] for s in summary['sheets']]
summary['derived']['total_rows'] = sum(s['non_empty_rows'] for s in summary['sheets'])

def sheet_non_empty(name):
    return next((s['non_empty_rows'] for s in summary['sheets'] if s['name'] == name), 0)

summary['derived']['content_inventory'] = {
    'learning_guides': sheet_non_empty('0_La Bàn Khởi Hành'),
    'marketing_templates': sheet_non_empty('13_Marketing 1 Chạm'),
    'disc_questions': sheet_non_empty('10_Trạm Đăng Kiểm Năng Lực'),
    'playbooks': sheet_non_empty('3_Bảo Bối Thực Chiến'),
    'empathy_language': sheet_non_empty('4_Ngôn Ngữ Thấu Cảm'),
    'underwriting_forms': sheet_non_empty('5_Trợ Lý Thẩm Định'),
    'leadership_lessons': sheet_non_empty('6_La Bàn Lãnh Đạo'),
    'rewards': sheet_non_empty('7_Trạm Tiếp Năng Lượng'),
    'points_transactions': sheet_non_empty('8_Ngân Hàng Điểm'),
    'feedback': sheet_non_empty('9_Góc Lắng Nghe'),
    'news_cases': sheet_non_empty('11_Bản Tin 90s & Án Lệ'),
    'daily_quizzes': sheet_non_empty('12_Nạp Não Mỗi Sáng'),
}

profile_rows = []
profile_ws = wb['1_Hồ Sơ Chiến Binh']
profile_headers = [safe_text(v) for v in next(profile_ws.iter_rows(values_only=True))]
for row in profile_ws.iter_rows(min_row=2, values_only=True):
    if any(v is not None and safe_text(v) for v in row):
        profile_rows.append(dict(zip(profile_headers, row)))

if profile_rows:
    streaks = [float(r.get('Chuỗi Ngày') or 0) for r in profile_rows]
    points = [float(r.get('Điểm') or 0) for r in profile_rows]
    summary['derived']['team_signals'] = {
        'profile_count': len(profile_rows),
        'total_xp': sum(points),
        'average_streak': round(sum(streaks) / len(streaks), 2),
        'max_streak': max(streaks),
        'hot_status_count': sum(1 for r in profile_rows if 'Hừng hực' in safe_text(r.get('Trạng Thái'))),
        'tired_status_count': sum(1 for r in profile_rows if 'mệt' in safe_text(r.get('Trạng Thái')).lower()),
        'disc_distribution': {key: sum(1 for r in profile_rows if safe_text(r.get('Nhóm DISC')) == key) for key in ['D', 'I', 'S', 'C']},
        'level_distribution': {key: sum(1 for r in profile_rows if safe_text(r.get('Cấp Độ App')) == key) for key in ['Rookie', 'Pro', 'Master']},
    }

log_ws = wb['2_Nhịp Đập Khách Hàng']
log_headers = [safe_text(v) for v in next(log_ws.iter_rows(values_only=True))]
log_rows = [dict(zip(log_headers, row)) for row in log_ws.iter_rows(min_row=2, values_only=True) if any(v is not None and safe_text(v) for v in row)]
if log_rows:
    xp_values = [float(r.get('Điểm Cộng') or 0) for r in log_rows]
    summary['derived']['customer_pulse_signals'] = {
        'log_count': len(log_rows),
        'total_xp_awarded': sum(xp_values),
        'average_service_score': round(sum(float(re.match(r'\d+', safe_text(r.get('Cấp Độ Dịch Vụ (1-6)'))).group()) for r in log_rows) / len(log_rows), 2),
        'wow_count': sum(1 for r in log_rows if 'WOW' in safe_text(r.get('Cấp Độ Dịch Vụ (1-6)'))),
        'public_log_count': sum(1 for r in log_rows if safe_text(r.get('Chế Độ Riêng Tư')) == 'Công khai'),
    }
OUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
print(json.dumps(summary['derived'], ensure_ascii=False, indent=2))
